import numpy as np
import math
from openpyxl import Workbook , load_workbook
min_cf = 0.05
y=0
s_cf_oned = np.empty((8760)) 
w_cf_oned = np.empty((8760)) 
l_cf_oned = np.empty((8760)) 
solar = np.zeros(1)

min_cp = np.empty(31)
Res = np.empty((365,24)) #再生能源總量
p_in = np.empty((365,24)) #多餘能源
p_out = np.empty((365,24)) #燃料電池發電
p_other = np.empty((365,24)) #其他能源
p_curl = np.empty((365,24)) #棄電
p_ele = np.empty((365,24))#電解槽上限
Electro_cp = np.zeros((8760,2))
Electro_cp2 = np.zeros((8760,2))
H_in = 0.65 #電解效率
H_out = 0.64 #燃料電池效率
BESS_in = 0.87 #電池效率
BESS_out = 0.87 #電池效率
solar_cp = 1400000 # 每kW太陽能版所需價格($/MW)
wind_cp = 4452000 # 每kW風力發電機所需價格($/MW)
eletro_cp = 1771000 # 每kW電解槽價格($/MW)
fuel_cp = 257000 # 每kW燃料電池價格($/MW)

for w in range(5): 
    #匯入solar
    wb_s_cf = load_workbook('C:\\Users\\farne\\Desktop\\專題\\表格\\SolarCF_2021.xlsx')
    sheet = wb_s_cf.worksheets[0]
    for i in range(1 ,sheet.max_row+1):
            s_cf_oned[i-1] = sheet.cell(row = i,column=1).value
            s_cf=s_cf_oned.reshape(365,24)
    #print (s_cf)   
    #匯入wind     
    wb_w_cf = load_workbook('C:\\Users\\farne\\Desktop\\專題\\表格\\OffshoreWind_CF_2022.xlsx')
    sheet = wb_w_cf.worksheets[0]
    for i in range(1 ,sheet.max_row+1):
            w_cf_oned[i-1] = sheet.cell(row = i,column=1).value
            w_cf=w_cf_oned.reshape(365,24)
    #匯入load     
    wb_l_cf = load_workbook('C:\\Users\\farne\\Desktop\\專題\\表格\\PowerGeneration_MainIsland.xlsx')
    sheet = wb_l_cf.worksheets[0]
    for i in range(1 ,sheet.max_row+1):
            l_cf_oned[i-1] = sheet.cell(row = i,column=1).value
            initial_load=l_cf_oned.reshape(365,24)
            load = initial_load + 8720
    wb = Workbook()
    ws1 = wb.create_sheet("1")
    ws2 = wb.create_sheet("2")
    ws3 = wb.create_sheet("3")
    ws4 = wb.create_sheet("4")
    #ws1 = wb["1"]
    solar_c = 70000 #太陽能裝置容量(MW)
    solar_fix = solar_c
    wind_c = 33000 #風能裝置容量(MW)
    x = 0
    t1 = 1
    e1 = np.zeros((8760)) #能量存儲
    while wind_c <=33000:
        x = 0#再生占比
        u=1#火力參數
        t1+=1
        if w < 7 :
            while  math.ceil(x*1000)/1000.0 != 0.60 :
                e1 = np.zeros(8760)
                solar_c = solar_fix
            
                while e1[8759] <= e1[0]:
                    if solar_c < 0 :
                        break
                    #初始變數
                    new_solar = solar_c*s_cf#新太陽算法
                    new_wind = wind_c*w_cf#新風算法
                    z=0
                    z1=0
                    solar_cost = solar_c * solar_cp  #太陽能板建置成本
                    wind_cost = wind_c * wind_cp #風力發電機建置成本
                    solar_maintain = solar_cost * 0.0208 # 每期太陽能維護費$
                    wind_maintain = wind_cost * 0.0324 # 每期風能維護費
                    Res_sum = 0
                    solar_sum = 0
                    wind_sum = 0
                    load_sum = 0
                    p_other_sum = 0
                    Rgen = 0
                    solar_gen = 0
                    wind_gen = 0
                    load_gen = 0
                    nRgen = 0
                    ele_t =0
                    i=0
                    j=0
                    k0 = 0
                    k1 = 0
                    k2 = 0
                    k3 = 0
                    k4 = 0
                    k5 = 0
                    kw = 0.062 # 火力每一度電幾美元 $/kW
                    co2 = 0.632085 #火火力排碳(kg/KWh)
                    co2_tax= 60
                    x=0
                    e1 =  e1 * 0

                    #能量平衡
                    l=0
                    Res = new_solar + new_wind
                    rl_i = load - Res #殘餘電荷
                    for i in range(365): #計算p_in
                        for j in range(24):
                            if rl_i[i][j] > 0:
                                p_in[i][j] = 0
                            else :
                                p_in[i][j] = -rl_i[i][j]
                    for i in range(365): #計算p_other
                        for j in range(24):
                            if rl_i[i][j] < 0:
                                p_other[i][j] = 0
                            else :
                                p_other[i][j] = u*rl_i[i][j]         
                    for i in range(365): #計算p_out
                        for j in range(24):
                            if rl_i[i][j] - p_other[i][j] <= 0:
                                p_out[i][j] = 0
                            else :
                                p_out[i][j] = rl_i[i][j] - p_other[i][j]
                    a = 0 
                    for i in range(365): #計算e1(MWh)
                            for j in range(24):
                                if i == 0 and j == 0 :
                                    e1[0] = 0 + (p_in[i][j] * 0.92 )- (p_out[i][j]/ 0.92)
                                else :
                                    e1[a] = e1[a-1]*0.999979 + (p_in[i][j] * 0.92 )- (p_out[i][j]/ 0.92)
                                a+=1
                                #print('e1[{}]={}'.format(a,e1[a]))
                    e1 = e1 + abs(np.min(e1))
                    
                    h_sto = (e1/33.3)*0.93 #計算h_sto (t)
                    #計算非再生能源總發電量
                    p_other_sum =  np.sum(p_other)
                    nRgen = np.sum(p_other_sum)*1000
                    #計算耗電量
                    load_sum =  np.sum(load)
                    load_gen = np.sum(load_sum)*1000
                    x = 1-(nRgen/load_gen)
                    if e1[8759] <= e1[0] :
                        solar_c = solar_c + 200
                if round(x,2) >=0.62:
                    u+=0.0001
                elif 0.62> round(x,3) >=0.603:
                    u+=0.0002
                elif 0.603> round(x,3) >=0.600:
                    u+=0.00009
                elif 0.59 < round(x,3) <= 0.597:
                    u-=0.003
                elif 0.597 < round(x,3) <= 0.599:
                    u-=0.0003
                elif round(x,2) <=0.59:
                    u-=0.008
            
                print(round(x,4))
            
            #成本計算
            #計算電解LCOS
            t = 20 #lifetime
            I_t = 1771000*np.max(p_in) #裝置成本
            O_t = 75*np.max(p_in) #維護費
            r = 0.05 #折現率
            E_t = 0
            #計算電解總發電量
            E_t = E_t + np.sum(p_in)
            E_gen = np.sum(E_t)*1000 #電解電度
            cost = I_t 
            gen = 0 
            for k1 in range(t): 
                cost = cost + (O_t)/((1+r)**k1)
            for q1 in range(t): 
                gen =  gen+(E_gen)/((1+r)**q1)
            electro_lcos = cost/(gen)

            electro_avecf = E_t/H_in/(np.max(p_in)*8760)#平均CF值
            
            #計算燃料電池LCOS
            t = 20 #lifetime
            I_t = fuel_cp*np.max(load) #裝置成本
            O_t = 1400*np.max(load) #維護費
            r = 0.05 #折現率
            E_t = 0
            #計算燃料電池總發電量
            E_t = E_t + np.sum(p_out)
            fuel_gen = np.sum(E_t)*1000 #電解電度
            cost = I_t 
            gen = 0 
            for k1 in range(t): 
                cost = cost + (O_t)/((1+r)**k1)
            for q2 in range(t): 
                gen =  gen+(fuel_gen)/((1+r)**q2)
            if gen == 0:
                fuel_lcos = 0
                fuel_cost = cost
            else:
                fuel_lcos = cost/(gen)
                fuel_cost = fuel_lcos * fuel_gen #燃料電池建置成本
            eletro_cost = electro_lcos * E_gen #電解槽建置成本
            
            storage_cost = ((277000/10)+6800) * np.max(e1) #儲氫建置成本
            
            #計算太陽能源總發電量
            solar_sum =  np.sum(new_solar)
            solar_gen = np.sum(solar_sum)*1000
            #計算風能源總發電量
            wind_sum =  np.sum(new_wind)
            wind_gen = np.sum(wind_sum)*1000
            #計算再生能源總發電量
            Res_sum =  np.sum(Res)
            Rgen = np.sum(Res_sum)*1000
            #計算非再生能源總發電量
            p_other_sum = np.sum(p_other)
            nRgen = np.sum(p_other_sum)*1000
            #計算耗電量
            load_sum = np.sum(load)
            load_gen = np.sum(load_sum)*1000
            x = 1-(nRgen/load_gen)

            #計算太陽能LCOE
            t = 25 #lifetime
            I_t = 43000/29.7*solar_c*1000 #裝置成本
            O_t = 1500/29.7*solar_c*1000 #+ 11*40000000 #維護費
            D_t = I_t * 0.00 #退役成本
            degrade = 0.000 #衰退率
            r = 0.05 #折現率
            cost = I_t 
            gen = 0
            for k1 in range(t): 
                cost = cost + (O_t)/((1+r)**k1)
            for k2 in range(t): 
                gen =  gen+(solar_gen)/((1+r)**k2)
            solar_lcoe = cost/(gen)
            #計算風能LCOS
            t = 20 #lifetime
            I_t = 154100/29.7*wind_c*1000 #裝置成本
            O_t = 4300/29.7*wind_c*1000 #維護費(美元)
            r = 0.05 #折現率
            cost = I_t 
            gen = 0
            for k1 in range(t): 
                cost = cost + (O_t)/((1+r)**(k1))
            for k2 in range(t): 
                gen =  gen+(wind_gen)/((1+r)**k2)
            wind_lcoe = cost/(gen)

            solar[t1-2]=solar_c
            
            min_cp[t1-2]=np.max(p_in)
            eletro_cost = 0
            tot_c = round((solar_gen*solar_lcoe+wind_gen*wind_lcoe+nRgen*kw+eletro_cost+fuel_cost+storage_cost+nRgen*co2*co2_tax/1000)/100000000,2)
            LCOE = (solar_gen*solar_lcoe+wind_gen*wind_lcoe)/(x*load_gen)
            LCOS = (eletro_cost+fuel_cost+storage_cost)/(x*load_gen)
            #print("再生能源佔比:" ,round(x,4))
            print("mincf為:",min_cf)
            print("太陽能、風能裝置容量:",solar_c,"MW",wind_c,"MW,占比:",x)
            print("成本:",tot_c,"億美元")
            print("存儲大小",round((np.max(h_sto)/1000),3),"千噸")
            print("電解槽容量:",round(np.max(p_in),2),"MW")
            print("存儲成本:" ,round(storage_cost/100000000,2),"億美元") 
            print("電解槽成本:" ,round(eletro_cost/100000000,2),"億美元")
            print("非再生能源每年發電:" ,round(nRgen/100000000,2),"億度，即",round(nRgen*kw/100000000,2),"億美元")
            print("太陽能每年成本:" ,solar_lcoe,",風能每年成本:" ,wind_lcoe,"再生每年成本",LCOE+LCOS)
            print("每年耗電:" ,round(load_gen/100000000,2),"億度",)
            """ print("風能每年發電:" ,round(wind_gen/100000000,2),"億度，即",round(wind_gen*wind_lcos/100000000,2),"億美元")
            print("電解LCOS為:",round(electro_lcos,3),"$/KWh")#LCOS    
            print("太陽能板成本:" ,solar_cost/100000000,"億美元, 每年太陽能板維護費:" ,solar_maintain/100000000,"億美元")
            print("風力發電機成本:" ,wind_cost/100000000,"億美元, 每年風能維護費:" ,wind_maintain/100000000,"億美元")        
            print("再生能源每年發電:" ,round(Rgen/100000000,2),"億度，即",round((solar_gen*solar_lcoe+wind_gen*wind_lcoe)/100000000,3),"億美元")
            print(round(solar_lcos,3),round(wind_lcos,3),round(electro_lcos,3),round(fuel_lcos,3))
            """
            ws1.cell(row =1,column=2,value ="太陽能")
            ws1.cell(row =1,column=3,value ="風能")
            ws1.cell(row = 1,column=4,value ="電解槽")
            ws1.cell(row = 1,column=5,value ="儲存空間")
            ws1.cell(row = 1,column=6,value ="總成本")
            ws1.cell(row = 1,column=7,value ="再生能源比")
            ws1.cell(row = 1,column=8,value ="碳排放(頓)")
            ws1.cell(row = 1,column=9,value ="碳排成本")
            ws1.cell(row = 1,column=10,value ="發電量")
            ws1.cell(row = 1,column=11,value ="平均cf")
            ws1.cell(row = 1,column=12,value ="lcoe")
            ws1.cell(row = 1,column=13,value ="lcoS")
            ws1.cell(row = 1,column=14,value ="20非再生lcoe")
            ws1.cell(row = 1,column=15,value ="60非再生lcoe")
            ws1.cell(row = 1,column=16,value ="100非再生lcoe")
            ws1.cell(row = 1,column=17,value ="整合lcoe")
            ws1.cell(row = 1,column=18,value ="整合lcoS")
            ws1.cell(row = 1,column=19,value ="混合20非再生lcoe")
            ws1.cell(row = 1,column=20,value ="混合60非再生lcoe")
            ws1.cell(row = 1,column=21,value ="混合100非再生lcoe")
            ws1.cell(row =t1,column=2,value =solar_c)
            ws1.cell(row =t1,column=3,value =wind_c)
            ws1.cell(row = t1,column=4,value =round(np.max(p_in),0))
            ws1.cell(row = t1,column=5,value =round((np.max(h_sto)/1000),2))
            ws1.cell(row = t1,column=6,value =tot_c)
            ws1.cell(row = t1,column=7,value =x)
            ws1.cell(row = t1,column=8,value = nRgen*co2/1000)
            ws1.cell(row = t1,column=9,value = nRgen*co2*co2_tax/1000)
            ws1.cell(row = t1,column=10,value = Rgen+nRgen)
            ws1.cell(row = t1,column=11,value = electro_avecf)
            ws1.cell(row = t1,column=12,value = LCOE)
            ws1.cell(row = t1,column=13,value = LCOS)
            ws1.cell(row = t1,column=14,value = ( nRgen*kw+nRgen*co2*20/1000)/((1-x)*load_gen))
            ws1.cell(row = t1,column=15,value = ( nRgen*kw+nRgen*co2*60/1000)/((1-x)*load_gen))
            ws1.cell(row = t1,column=16,value = ( nRgen*kw+nRgen*co2*100/1000)/((1-x)*load_gen))
            LCOE = (solar_gen*solar_lcoe+wind_gen*wind_lcoe)/(load_gen)
            LCOS = (eletro_cost+fuel_cost+storage_cost)/(load_gen)
            ws1.cell(row = t1,column=17,value = LCOE)
            ws1.cell(row = t1,column=18,value = LCOS)
            ws1.cell(row = t1,column=19,value = (nRgen*kw+nRgen*co2*20/1000)/(load_gen))
            ws1.cell(row = t1,column=20,value = (nRgen*kw+nRgen*co2*60/1000)/(load_gen))
            ws1.cell(row = t1,column=21,value = (nRgen*kw+nRgen*co2*100/1000)/(load_gen))
        
        
        Smin_cf  =str(round(min_cf*100,1))    
        
        if wind_c <= 33000 :
            wind_c = wind_c + 1000
            e1 = np.zeros((365,25))
            solar_c = solar_c - 10000
            solar_fix = solar_c
        if solar_c<=0:
            break
        print(wind_c)
        Smin_cf  =str(round(min_cf*100,1)) 
        wb.save('C:\\Users\\farne\\Desktop\\專題\\code\\Li24hrNR'+Smin_cf+'%min_cf.xlsx')
    min_cf+=0.05
    