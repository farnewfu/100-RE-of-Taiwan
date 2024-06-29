import numpy as np
import math
from openpyxl import Workbook , load_workbook
min_cf = 0.05
y=0
s_cf_oned = np.empty((8760)) 
w_cf_oned = np.empty((8760)) 
l_cf_oned = np.empty((8760)) 
solar = np.zeros(1)

for w in range(5): 
    #匯入solar
    wb_s_cf = load_workbook('C:\\Users\\farne\\Desktop\\專題\\表格\\SolarCF_2021.xlsx')
    sheet = wb_s_cf.worksheets[0]
    for i in range(1 ,sheet.max_row+1):
            s_cf_oned[i-1] = sheet.cell(row = i,column=1).value
            s_cf=s_cf_oned.reshape(365,24)
    #print (s_cf)   
    #匯入wind     
    wb_w_cf = load_workbook('C:\\Users\\farne\\Desktop\\專題\\表格\\OffshoreWind_CF_2022.xlsx')
    sheet = wb_w_cf.worksheets[0]
    for i in range(1 ,sheet.max_row+1):
            w_cf_oned[i-1] = sheet.cell(row = i,column=1).value
            w_cf=w_cf_oned.reshape(365,24)
    #匯入load     
    wb_l_cf = load_workbook('C:\\Users\\farne\\Desktop\\專題\\表格\\PowerGeneration_MainIsland.xlsx')
    sheet = wb_l_cf.worksheets[0]
    for i in range(1 ,sheet.max_row+1):
            l_cf_oned[i-1] = sheet.cell(row = i,column=1).value
            initial_load=l_cf_oned.reshape(365,24)
            load = initial_load + 8720
    wb = Workbook()
    ws1 = wb.create_sheet("1")
    ws2 = wb.create_sheet("2")
    ws3 = wb.create_sheet("3")
    ws4 = wb.create_sheet("4")
    #ws1 = wb["1"]
    solar_c = 150000 #太陽能裝置容量(MW)
    wind_c = 50000 #風能裝置容量(MW)
    
    x = 0
    t1 = 1
    e1 = np.zeros((8760)) #能量存儲
    while wind_c <=50000:
        t1+=1
    
        while e1[8759] <= e1[0]:
            if solar_c < 0 :
                break
            #初始變數
            new_solar = solar_c*s_cf#新太陽算法
            new_wind = wind_c*w_cf#新風算法
            z=0
            z1=0
            load_o = np.array([25400,24300,23500,22900,22700,22700,23100,24000,27000,30000,30500,30500,30500
                            ,28500,30500,30500,31000,31000,30500,30500,30000,29000,28000,26780])+2000 #一天的負載(MW)
            load_d = [860,-1148,-1778,-1386,3297,3242,5204,4235,4967,1894,-867,-1723] #每月負載變化量
            month_d = np.array([31,28,31,30,31,30,31,31,30,31,30,31])#每月天數
            n=1
            Res = np.empty((365,24)) #再生能源總量
            p_in = np.empty((365,24)) #多餘能源
            p_out = np.empty((365,24)) #燃料電池發電
            p_other = np.empty((365,24)) #其他能源
            p_curl = np.empty((365,24)) #棄電
            p_ele = np.empty((365,24))#電解槽上限
            Electro_cp = np.zeros((8760,2))
            Electro_cp2 = np.zeros((8760,2))
            solar_cf = np.array([0.222,0.243,0.297,0.359,0.399,0.339,0.402,0.318,0.379,0.404,0.314,0.258])
            solar_all =  solar_c * solar_cf #12個月太陽能變化
            solar_d = np.array([0,0,0,0,0,0,0,0,0.5,1,1,1,1,1,1,1,1,1,0.5,0,0,0,0,0]) #一天太陽能變化
            wind_d =  np.array([1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1])#一天風能變化
            t_step  = np.array([1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1]) #時間步長度(hr)
            
            #wind_cf = np.array([0.407581205,0.297847618,0.272090101,0.275849813,0.108940136,0.138544551,
                                #0.145178875,0.079656181,0.077650467,0.37225009,0.505605805,0.660649456])+0.12 #風能每月容量因數+離岸
            wind_cf = np.array([0.617614713,0.386152457,0.275921341,0.320284422,0.286013465,0.248966567,
            0.243053504,0.240173943,0.179302256,0.361230468,0.459894779,0.612115236])#offshore風能每月容量因數
            wind = wind_cf * wind_c #12個月風能變化
            H_in = 0.65 #電解效率
            H_out = 0.64 #燃料電池效率
            BESS_in = 0.87 #電池效率
            BESS_out = 0.87 #電池效率
            solar_cp = 1400000 # 每kW太陽能版所需價格($/kW)
            wind_cp = 4452000 # 每kW風力發電機所需價格($/kW)
            eletro_cp = 1771000 # 每kW電解槽價格($/kW)
            fuel_cp = 1320000 # 每kW燃料電池價格($/kW)
            solar_cost = solar_c * solar_cp  #太陽能板建置成本
            wind_cost = wind_c * wind_cp #風力發電機建置成本
            solar_maintain = solar_cost * 0.0208 # 每期太陽能維護費$
            wind_maintain = wind_cost * 0.0324 # 每期風能維護費$
            Res_sum = 0
            solar_sum = 0
            wind_sum = 0
            load_sum = 0
            p_other_sum = 0
            Rgen = 0
            solar_gen = 0
            wind_gen = 0
            load_gen = 0
            nRgen = 0
            ele_t =0
            i=0
            j=0
            k0 = 0
            k1 = 0
            k2 = 0
            k3 = 0
            k4 = 0
            k5 = 0
            kw = 0.067 # 火力每一度電幾美元 $/kW
            co2 = 0.632085 #火火力排碳(kg/KWh)
            x=0

            #能量平衡
            l=0
            Res = new_solar + new_wind
            rl_i = load - Res #殘餘電荷
            """for i in range(365):
                for k in range (24):
                    print(rl_i[i][k])"""
            
            for i in range(365): #計算p_in
                for j in range(24):
                    if rl_i[i][j] > 0:
                        p_in[i][j] = 0
                    else :
                        p_in[i][j] = -rl_i[i][j]
            for i in range(365): #計算p_other
                for j in range(24):
                    if rl_i[i][j] < 0:
                        p_other[i][j] = 0
                    else :
                        #p_other[i][j] = rl_i[i][j]
                        p_other[i][j] = 0
            for i in range(365): #計算p_out
                for j in range(24):
                    if rl_i[i][j] - p_other[i][j] <= 0:
                        p_out[i][j] = 0
                    else :
                        p_out[i][j] = rl_i[i][j] - p_other[i][j]
            a = 0 
            for i in range(365): #計算e1(MWh)
                    for j in range(24):
                        if i == 0 and j == 0 :
                            e1[0] = 0 + (p_in[i][j] * H_in )- (p_out[i][j]/ H_out)
                        else :
                            e1[a] = e1[a-1]+ (p_in[i][j] * H_in )- (p_out[i][j]/ H_out)
                        a+=1
                        #print('e1[{}][{}]={}'.format(a,j,e1[a][j]))
            
            e1 = e1 + abs(np.min(e1))
            
            h_sto = (e1/33.3)*0.93 #計算h_sto (t)
            if e1[8759] <= e1[0] :
                solar_c = solar_c + 200
        """if wind_c == 60000 or wind_c == 65000 or wind_c == 70000 or wind_c == 75000:       
            print(solar[0],solar[5])   
            print(wind[0],wind[5]) """
        
        #成本計算
        #計算電解LCOS
        t = 20 #lifetime
        I_t = 1771000*np.max(p_in) #裝置成本
        O_t = 75*np.max(p_in) #維護費
        r = 0.05 #折現率
        E_t = 0
        #計算電解總發電量
        E_t = E_t + np.sum(p_in)
        E_gen = np.sum(E_t)*1000 #電解電度
        cost = I_t 
        gen = 0 
        for k1 in range(t): 
            cost = cost + (O_t)/((1+r)**k1)
        for q1 in range(t): 
            gen =  gen+(E_gen)/((1+r)**q1)
        electro_lcos = cost/(gen)

        electro_avecf = E_t/H_in/(np.max(p_in)*8760)#平均CF值
        
        #計算燃料電池LCOS
        t = 20 #lifetime
        I_t = 1333000*np.max(load) #裝置成本
        O_t = 12.7*np.max(load) #維護費
        r = 0.05 #折現率
        E_t = 0
        #計算燃料電池總發電量
        E_t = E_t + np.sum(p_out)
        fuel_gen = np.sum(E_t)*1000 #電解電度
        cost = I_t 
        gen = 0 
        for k1 in range(t): 
            cost = cost + (O_t)/((1+r)**k1)
        for q2 in range(t): 
            gen =  gen+(fuel_gen)/((1+r)**q2)
        if gen == 0:
            fuel_lcos = 0
            fuel_cost = cost
        else:
            fuel_lcos = cost/(gen)
            fuel_cost = fuel_lcos * fuel_gen #燃料電池建置成本
        eletro_cost = electro_lcos * E_gen #電解槽建置成本
        storage_cost = ((1179/20)+29) * np.max(e1) #儲氫建置成本
        
        #計算太陽能源總發電量
        solar_sum = solar_sum + np.sum(new_solar)
        solar_gen = np.sum(solar_sum)*1000
        #計算風能源總發電量
        wind_sum = wind_sum + np.sum(new_wind)
        wind_gen = np.sum(wind_sum)*1000
        #計算再生能源總發電量
        Res_sum = Res_sum + np.sum(Res)
        Rgen = np.sum(Res_sum)*1000
        #計算非再生能源總發電量
        p_other_sum = p_other_sum + np.sum(p_other)
        nRgen = np.sum(p_other_sum)*1000
        #計算耗電量
        load_sum = load_sum + np.sum(load)
        load_gen = np.sum(load_sum)*1000
        x = 1-(nRgen/load_gen)

        #計算太陽能LCOE
        t = 25 #lifetime
        I_t = 43000/29.7*solar_c*1000 #裝置成本
        O_t = 1500/29.7*solar_c*1000 #+ 11*40000000 #維護費
        D_t = I_t * 0.00 #退役成本
        degrade = 0.000 #衰退率
        r = 0.05 #折現率
        cost = I_t 
        gen = 0
        for k1 in range(t): 
            cost = cost + (O_t)/((1+r)**k1)
        for k2 in range(t): 
            gen =  gen+(solar_gen)/((1+r)**k2)
        solar_lcoe = cost/(gen)
        #計算風能LCOS
        t = 20 #lifetime
        I_t = 154100/29.7*wind_c*1000 #裝置成本
        O_t = 4300/29.7*wind_c*1000 #維護費(美元)
        r = 0.05 #折現率
        cost = I_t 
        gen = 0
        for k1 in range(t): 
            cost = cost + (O_t)/((1+r)**(k1))
        for k2 in range(t): 
            gen =  gen+(wind_gen)/((1+r)**k2)
        wind_lcoe = cost/(gen)
        solar[t1-2]=solar_c
        
        tot_c = round((solar_gen*solar_lcoe+wind_gen*wind_lcoe+nRgen*kw+eletro_cost+fuel_cost+storage_cost)/100000000,2)
        LCOE = (solar_gen*solar_lcoe+wind_gen*wind_lcoe)/(x*load_gen)
        LCOS = (eletro_cost+fuel_cost+storage_cost)/(x*load_gen)
        print("mincf為:",min_cf)
        print("太陽能、風能裝置容量:",solar_c,"MW",wind_c,"MW")
        print("成本:",tot_c,"億美元")
        print("存儲大小",round((np.max(h_sto)/1000),3),"千噸")
        print("電解槽容量:",round(np.max(p_in),2),"MW")
        print("燃料電池成本:" ,round(fuel_cost/100000000,2),"億美元")
        print("存儲成本:" ,round(storage_cost/100000000,2),"億美元") 
        print("電解槽成本:" ,round(eletro_cost/100000000,2),"億美元")
        print("太陽能、風能成本:",solar_gen*solar_lcoe/100000000,"億美元",wind_gen*wind_lcoe/100000000,"億美元")
        print("太陽能每年發電:" ,round(solar_gen/100000000,2),"億度，即",solar_lcoe,"美元")
        print("風能每年發電:" ,round(wind_gen/100000000,2),"億度，即",wind_lcoe,"美元") 
        print("每年耗電:" ,round(load_gen/100000000,2),"億度")
        """print("電解LCOS為:",round(electro_lcos,3),"$/KWh")#LCOS    
        print("太陽能板成本:" ,solar_cost/100000000,"億美元, 每年太陽能板維護費:" ,solar_maintain/100000000,"億美元")
        print("風力發電機成本:" ,wind_cost/100000000,"億美元, 每年風能維護費:" ,wind_maintain/100000000,"億美元")        
        print("再生能源每年發電:" ,round(Rgen/100000000,2),"億度，即",round((solar_gen*solar_lcoe+wind_gen*wind_lcoe)/100000000,3),"億美元")
        print("非再生能源每年發電:" ,round(nRgen/100000000,2),"億度，即",round(nRgen*kw/100000000,2),"億美元")
        print(round(solar_lcos,3),round(wind_lcos,3),round(electro_lcos,3),round(fuel_lcos,3))
    """
        ws1.cell(row =1,column=2,value ="太陽能")
        ws1.cell(row =1,column=3,value ="風能")
        ws1.cell(row = 1,column=4,value ="電解槽")
        ws1.cell(row = 1,column=5,value ="儲存空間")
        ws1.cell(row = 1,column=6,value ="總成本")
        ws1.cell(row = 1,column=7,value ="再生能源比")
        ws1.cell(row = 1,column=8,value ="發電量")
        ws1.cell(row = 1,column=9,value ="平均cf")
        ws1.cell(row = 1,column=10,value ="LCOE")
        ws1.cell(row = 1,column=11,value ="LCOS")
        ws1.cell(row =1,column=12,value ="太陽能電")
        ws1.cell(row =1,column=13,value ="風能電")
        ws1.cell(row =t1,column=2,value =solar_c)
        ws1.cell(row =t1,column=3,value =wind_c)
        ws1.cell(row = t1,column=4,value =round(np.max(p_in),0))
        ws1.cell(row = t1,column=5,value =round((np.max(h_sto)/1000),2))
        ws1.cell(row = t1,column=6,value =tot_c)
        ws1.cell(row = t1,column=7,value =x)
        ws1.cell(row = t1,column=8,value = Rgen+nRgen)
        ws1.cell(row = t1,column=9,value = electro_avecf)
        ws1.cell(row = t1,column=10,value = LCOE)
        ws1.cell(row = t1,column=11,value = LCOS)
        ws1.cell(row =t1,column=12,value =solar_gen/100000000)
        ws1.cell(row =t1,column=13,value =wind_gen/100000000)
        
        if wind_c == 60000 or wind_c == 70000 or wind_c == 80000:
            for t in range(8760):
                if wind_c == 60000:
                    t2=2
                elif wind_c == 70000:
                    t2=3
                else :
                    t2=4
                #ws2.cell(row=t+1,column=t2*2+3,value=Electro_order[t][0])
                #ws2.cell(row=t+1,column=t2*2+4,value=Electro_order[t][1])

        
        #計算min_cf用
        if wind_c != 0 :#or wind_c == 70000 or wind_c == 74000 or wind_c == 80000 :
            b=0
            p=0
            a=0
            b=0
            g = 0 
            count = 0
            count2 = 0
            b= 0
            for p1 in range(365):
                    for p2 in range(24):
                        b=1   
                        Electro_cp[a][0] = p_in[p1][p2] 
                        Electro_cp[a][1] = b
                        a+=1
            Electro_order = sorted(Electro_cp,key=lambda Electro_cp:Electro_cp[0])#由小到大排矩陣
            
            
        Electro_mincp = np.max(p_in) #min_cp值    
        Electro_mincf=0
        solar_c_2 = solar[t1-2]
        
        while abs (np.round(Electro_mincf,3) - min_cf) > 0.01*min_cf:
            
            if 0.02 <= min_cf- Electro_mincf and min_cf- Electro_mincf<0.05:
                Electro_mincp -= 200
            elif 0.00 <= min_cf- Electro_mincf and min_cf- Electro_mincf<0.02:
                Electro_mincp -= 50
            elif 0.05 <= min_cf- Electro_mincf and min_cf- Electro_mincf<0.1:
                Electro_mincp -= 500
            elif 0.1 <= min_cf- Electro_mincf :
                Electro_mincp -= 1000
            elif Electro_mincf>= min_cf:
                Electro_mincp += 50
            else : 
                Electro_mincp -= 50
            e1 = np.zeros((8760))#重新計算p_in,e1
            
            while e1[8759] <= e1[0]:
                new_solar =  solar_c_2 * s_cf
                Res = new_solar + new_wind
                rl_i = load - Res #殘餘電荷
                i=0 #初始化
                j=0 #初始化
                k=0
                a=0
                for i in range(365): #計算新p_in
                    for j in range(24):
                        #print('p_in[{}][{}]={}'.format(i,j,p_in[i][j]))
                        if rl_i[i][j] > 0:
                            p_in[i][j] = 0
                            p_curl[i][j] = 0
                        elif -rl_i[i][j] < Electro_mincp:
                            p_in[i][j] = -rl_i[i][j]
                            p_curl[i][j] = 0
                        elif -rl_i[i][j] > Electro_mincp:
                            p_in[i][j] = Electro_mincp       
                            p_curl[i][j] = -rl_i[i][j] - Electro_mincp
                for i in range(365): #計算e1(MWh)
                        for j in range(24):
                            if a == 0 and j == 0 :
                                e1[a] = 0 + (p_in[i][j] * H_in )- (p_out[i][j]/ H_out)
                            else :
                                e1[a] = e1[a-1] + (p_in[i][j] * H_in )- (p_out[i][j]/ H_out)
                         
                            #print('e1[{}][{}]={}'.format(a,j,e1[a][j]))
                            a += 1
                e1 = e1 + abs(np.min(e1))
        
                h_sto2 = (e1/33.3)*0.93 #計算h_sto (t)
            
                
                b=0
                p=0
                a=0
                b=0
                g = 0 
                count = 0
                count2=0
                Electro_gen = 0
                b= 0
                for p1 in range(365):
                        for p2 in range(24):
                            b=1   
                            Electro_cp2[a][0] = p_in[p1][p2] 
                            Electro_cp2[a][1] = b
                            a+=1
                Electro_order2 = sorted(Electro_cp2,key=lambda Electro_cp2:Electro_cp2[0])#由小到大排矩陣   
                while (count/8760) < 1 :#計算滿發電時間g
                    count = count + Electro_order2[g][1]
                    if Electro_order2[g][0] == Electro_mincp :
                        count2 = count2 + Electro_order2[g][1]
                    g+=1
                    Electro_mincf=count2/8760
                print(Electro_mincf,Electro_mincp,solar_c_2,min_cf- Electro_mincf)
                if e1[8759] <= e1[0]:
                    if min_cf- Electro_mincf > 0 :
                        solar_c_2 = solar_c_2 + 100
                    else:
                         break\
                         
                    #elif min_cf- Electro_mincf < 0.05 and min_cf- Electro_mincf > 0 :
                       # solar_c_2 = solar_c_2 + 100
                    """elif min_cf- Electro_mincf > -0.1 and min_cf- Electro_mincf < -0.05 :
                        solar_c_2 = solar_c_2 - 500
                    elif min_cf- Electro_mincf > -0.05 and min_cf- Electro_mincf < 0 :
                        solar_c_2 = solar_c_2 - 100"""
        
        print(Electro_mincf,Electro_mincp,solar_c_2)
        #計算p_curl
        E_t = 0 
        E_t =  np.sum(p_curl)
        p_curl_tot = np.sum(E_t)*1000 #電度   


        #計算新電解LCOS
        t = 20 #lifetime
        I_t = 1771000*Electro_mincp #裝置成本
        O_t = 75*Electro_mincp #維護費
        r = 0.05 #折現率
        E_t = 0 #計算新電解量

        E_t =  np.sum(p_in)
        E_gen2 = np.sum(E_t)*1000 #電解電度
        cost = I_t 
        gen = 0 
        for k1 in range(t): 
            cost = cost + (O_t)/((1+r)**k1)
        for q1 in range(t): 
            gen =  gen+(E_gen2)/((1+r)**q1)
        electro_lcos2 = cost/(gen)

        eletro_cost = electro_lcos2 * E_gen2 #電解槽建置成本
        fuel_cost = fuel_lcos * fuel_gen #燃料電池建置成本
        storage_cost = ((1179/20)+29) * np.max(e1) #儲氫建置成本
        avg_cf = (E_t/H_in)/(Electro_mincp*8760) 
        solar_sum = 0
        wind_sum=0
        Res_sum=0
        p_other_sum=0
        load_sum=0
        #計算太陽能源總發電量
        solar_sum = solar_sum + np.sum(new_solar)
        solar_gen = np.sum(solar_sum)*1000
        #計算風能源總發電量
        wind_sum = wind_sum + np.sum(new_wind)
        wind_gen = np.sum(wind_sum)*1000
        #計算再生能源總發電量
        Res_sum = Res_sum + np.sum(Res)
        Rgen = np.sum(Res_sum)*1000
        #計算非再生能源總發電量
        p_other_sum = p_other_sum + np.sum(p_other)
        nRgen = np.sum(p_other_sum)*1000
        #計算耗電量
        load_sum = load_sum + np.sum(load)
        load_gen = np.sum(load_sum)*1000
        x = 1-(nRgen/load_gen)

        #計算新太陽能LCOE
        t = 25 #lifetime
        I_t = 43000/29.7*solar_c_2*1000 #裝置成本
        O_t = 1500/29.7*solar_c_2*1000 #+ 11*40000000 #維護費
        D_t = I_t * 0.00 #退役成本
        degrade = 0.000 #衰退率
        r = 0.05 #折現率
        cost = I_t 
        gen = 0
        for k1 in range(t): 
            cost = cost + (O_t)/((1+r)**k1)
        for k2 in range(t): 
            gen =  gen+(solar_gen)/((1+r)**k2)
        solar_lcoe2 = cost/(gen)

        tot_c = round((solar_gen*solar_lcoe2+wind_gen*wind_lcoe+nRgen*kw+eletro_cost+fuel_cost+storage_cost)/100000000,2)
        LCOE = (solar_gen*solar_lcoe2+wind_gen*wind_lcoe)/(x*load_gen)
        LCOS = (eletro_cost+fuel_cost+storage_cost)/(x*load_gen)
        LCOC = LCOE * p_curl_tot/load_gen
        print("mincf為:",min_cf)
        print("新太陽能、風能裝置容量:",solar_c_2,"MW",wind_c,"MW,占比:",x)
        print("新成本:",tot_c,"億美元")
        print("新存儲大小",round((np.max(h_sto2)/1000),3),"千噸")
        print("新電解槽容量:",round(Electro_mincp,2),"MW")
        print("新燃料電池成本:" ,round(fuel_cost/100000000,2),"億美元")
        print("新存儲成本:" ,round(storage_cost/100000000,2),"億美元") 
        print("新電解槽成本:" ,round(eletro_cost/100000000,2),"億美元")
        print("新太陽能、風能成本:",solar_gen*solar_lcoe/100000000,"億美元",wind_gen*wind_lcoe/100000000,"億美元")
        print("新非再生能源每年發電:" ,round(nRgen/100000000,2),"億度，即",round(nRgen*kw/100000000,2),"億美元")
        print("新太陽能、風能LCOE:",solar_lcoe2,"美元",wind_lcoe,"美元")
        #print(electro_avecf,avg_cf)
        ws2.cell(row =1,column=2,value ="太陽能")
        ws2.cell(row =1,column=3,value ="風能")
        ws2.cell(row = 1,column=4,value ="電解槽")
        ws2.cell(row = 1,column=5,value ="儲存空間")
        ws2.cell(row = 1,column=6,value ="總成本")
        ws2.cell(row = 1,column=7,value ="再生能源比")
        ws2.cell(row = 1,column=8,value ="棄電(度)")
        ws2.cell(row = 1,column=9,value ="發電量")
        ws2.cell(row = 1,column=10,value ="平均cf")
        ws2.cell(row = 1,column=11,value ="總再生lcoe")
        ws2.cell(row = 1,column=12,value ="1級lcoe")
        ws2.cell(row = 1,column=13,value ="LCOS")
        ws2.cell(row = 1,column=14,value ="LCOC")
        ws2.cell(row = t1,column=2,value = solar_c_2)
        ws2.cell(row = t1,column=3,value = wind_c)
        ws2.cell(row = t1,column=4,value = Electro_mincp)
        ws2.cell(row = t1,column=5,value = round((np.max(h_sto2)/1000),3))
        ws2.cell(row = t1,column=6,value = tot_c)
        ws2.cell(row = t1,column=7,value =x)
        ws2.cell(row = t1,column=8,value = p_curl_tot)
        ws2.cell(row = t1,column=9,value = Rgen+nRgen)
        ws2.cell(row = t1,column=10,value = avg_cf)
        ws2.cell(row = t1,column=11,value = LCOE+LCOS+LCOC)
        ws2.cell(row = t1,column=12,value = LCOE)
        ws2.cell(row = t1,column=13,value = LCOS)
        ws2.cell(row = t1,column=14,value = LCOC)
        
        a=0
        b=0
        c=0
        
        if wind_c == 50000 :
            b=1
        elif wind_c == 70000:
            b=6
        elif wind_c == 75000:
            b=11
        if wind_c == 50000 or wind_c == 70000 or wind_c == 75000:
            for z1 in range(365):
                    for z2 in range(24):
                        ws3.cell(row = a+1, column=b, value=h_sto[a] )
                        a+=1
            a=0 
            for z1 in range(365):
                    for z2 in range(24):
                        ws3.cell(row = a+1, column=b+1, value= Electro_order[a][0])
                        ws3.cell(row = a+1, column=b+2, value= Electro_order[a][1])
                        a+=1
            a=0
            for z1 in range(365):
                    for z2 in range(24):
                        ws3.cell(row = a+1, column=b+3, value=Electro_order2[a][0])
                        ws3.cell(row = a+1, column=b+4, value=Electro_order2[a][1])
                        a+=1
            a=0
            for z1 in range(365):
                    for z2 in range(24):
                        ws4.cell(row = z2+1, column=z1+1, value=rl_i[z1][z2])
                        a+=1

        Smin_cf  =str(round(min_cf*100,1))    
        wb.save('C:\\Users\\farne\\Desktop\\專題\\code\\cf24hrR'+Smin_cf+'%min_cf.xlsx')
        if wind_c <= 50000 :
            wind_c = wind_c + 1000
            e1 = np.zeros(8760)
            solar_c = solar_c - 10000
        print(load_sum)
    min_cf+=0.05

        